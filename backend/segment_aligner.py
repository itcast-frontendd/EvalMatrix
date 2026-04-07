"""
segment_aligner.py — 同传翻译多模型分句对齐工具

核心思路：
1. 以人工复核的 ASR 原文为基准
2. 用句号/问号/感叹号等终止符将原文切分为「语义段落」
3. 将各模型的译文按原文文本匹配映射到对应段落
4. 输出：每行一个语义段落，包含原文 + 各模型译文（已合并对齐）

用法：
  python segment_aligner.py --origin <原文.txt> --model1 <模型1.xlsx> --model2 <模型2.xlsx> --output <对齐结果.xlsx>
  
也可作为 API 被后端调用。
"""

import re
import argparse
import json
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass, field

import openpyxl
import pandas as pd


# ── 数据结构 ──

@dataclass
class SourceChunk:
    """原文的一个 ASR chunk（一行）"""
    idx: int
    text: str

@dataclass 
class ModelRow:
    """模型输出的一行"""
    idx: int
    asr_text: str      # B 列：原文(STT)
    translation: str   # C 列：翻译
    extra: dict = field(default_factory=dict)  # D-H 列的延迟数据等

@dataclass
class AlignedSegment:
    """对齐后的语义段落"""
    segment_id: int
    source_text: str           # 合并后的原文
    source_chunks: List[int]   # 原始 chunk 行号
    models: Dict[str, str]     # {model_name: 合并后的翻译}
    models_asr: Dict[str, str] # {model_name: 该模型看到的 ASR 原文}
    models_rows: Dict[str, List[int]]  # {model_name: 对应的原始行号}


# ── 1. 读取数据 ──

def read_origin_txt(path: str) -> List[SourceChunk]:
    """读取人工复核过的 ASR 原文 TXT"""
    chunks = []
    with open(path, "r", encoding="utf-8") as f:
        for i, line in enumerate(f):
            text = line.rstrip("\n\r")
            if text.strip():
                chunks.append(SourceChunk(idx=i, text=text))
    return chunks


def read_model_excel(path: str, asr_col: int = 2, trans_col: int = 3) -> List[ModelRow]:
    """读取模型 Excel (A=ID, B=原文, C=翻译, D-H=延迟)"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):  # skip header
        asr = ws.cell(r, asr_col).value
        trans = ws.cell(r, trans_col).value
        if asr is None and trans is None:
            continue
        extra = {}
        for c in range(4, min(ws.max_column + 1, 9)):
            header = ws.cell(1, c).value
            extra[str(header or f"col_{c}")] = ws.cell(r, c).value
        rows.append(ModelRow(
            idx=r - 2,  # 0-based
            asr_text=str(asr or "").strip(),
            translation=str(trans or "").strip(),
            extra=extra,
        ))
    return rows


# ── 2. 语义段落切分 ──

# 句子结束符模式
SENTENCE_END_PATTERN = re.compile(
    r'[.!?。！？…]+[\s"\'）\)]*$'  # 英文/中文句号、问号、感叹号、省略号
    r'|'
    r'[.!?。！？…]+[\s"\'）\)]*(?=\s)',  # 句末标点后跟空格
    re.UNICODE
)

def split_into_segments(chunks: List[SourceChunk], 
                        min_segment_chars: int = 20,
                        max_segment_chars: int = 500) -> List[List[int]]:
    """
    将原文 chunks 按句子终止符切分为语义段落。
    
    策略：
    - 扫描每个 chunk，累积文本
    - 当遇到句末标点（.!?）且累积长度 >= min_segment_chars，切分
    - 当累积长度 > max_segment_chars，强制切分
    - 返回: [[chunk_idx, ...], [chunk_idx, ...], ...]
    """
    segments: List[List[int]] = []
    current_group: List[int] = []
    current_text = ""
    
    for chunk in chunks:
        current_group.append(chunk.idx)
        current_text += " " + chunk.text
        
        text_stripped = current_text.strip()
        
        # 检查是否在句子结束处
        ends_with_sentence = bool(re.search(r'[.!?。！？…][\s"\'）\)]*$', text_stripped))
        
        should_split = False
        if ends_with_sentence and len(text_stripped) >= min_segment_chars:
            should_split = True
        elif len(text_stripped) > max_segment_chars:
            should_split = True
        
        if should_split:
            segments.append(current_group)
            current_group = []
            current_text = ""
    
    # 剩余的 chunks
    if current_group:
        if segments:
            # 如果最后一组太短，合并到前一组
            if len(current_text.strip()) < min_segment_chars // 2:
                segments[-1].extend(current_group)
            else:
                segments.append(current_group)
        else:
            segments.append(current_group)
    
    return segments


# ── 3. 模型译文对齐 ──

def normalize_text(text: str) -> str:
    """归一化文本用于匹配"""
    return re.sub(r'\s+', ' ', text.lower().strip())


def align_model_to_segments(
    model_rows: List[ModelRow],
    source_chunks: List[SourceChunk],
    segments: List[List[int]],
) -> List[Tuple[List[int], str, str]]:
    """
    将模型的输出行映射到语义段落。
    
    策略：贪心字符偏移匹配
    - 将原文 chunks 拼接成全文，记录每个 chunk 的字符区间
    - 将模型的 ASR 文本也拼接，记录每行的字符区间  
    - 通过字符偏移找到模型每行对应的原文 chunk 范围
    - 再从 chunk 范围映射到 segment
    
    Returns: [(model_row_indices, merged_asr, merged_translation), ...] per segment
    """
    
    # Build source full text with chunk boundaries
    source_full = ""
    chunk_spans = {}  # chunk_idx -> (start, end)
    for chunk in source_chunks:
        start = len(source_full)
        source_full += chunk.text + " "
        end = len(source_full)
        chunk_spans[chunk.idx] = (start, end)
    source_full_norm = normalize_text(source_full)
    
    # Build model full text and find each row's position in source
    model_to_chunks: Dict[int, List[int]] = {}  # model_row_idx -> [chunk_idx, ...]
    
    # Strategy: sequential matching via cumulative text
    # Model ASR texts should be subsequences of source full text
    search_start = 0
    for mrow in model_rows:
        mtext_norm = normalize_text(mrow.asr_text)
        if not mtext_norm:
            continue
        
        # Find this model's ASR text in the source full text
        pos = source_full_norm.find(mtext_norm, max(0, search_start - 50))
        if pos == -1:
            # Fuzzy fallback: try matching first 30 chars
            prefix = mtext_norm[:30]
            pos = source_full_norm.find(prefix, max(0, search_start - 50))
        if pos == -1:
            # Last resort: just advance sequentially
            pos = search_start
        
        match_end = pos + len(mtext_norm)
        search_start = match_end
        
        # Find which source chunks this span covers
        matched_chunks = []
        for chunk in source_chunks:
            cs, ce = chunk_spans[chunk.idx]
            # Overlap check
            if cs < match_end and ce > pos:
                matched_chunks.append(chunk.idx)
        
        if not matched_chunks:
            # Fallback: assign to nearest chunk by position
            for chunk in source_chunks:
                cs, ce = chunk_spans[chunk.idx]
                if ce >= pos:
                    matched_chunks = [chunk.idx]
                    break
        
        model_to_chunks[mrow.idx] = matched_chunks
    
    # Build chunk_to_segment mapping
    chunk_to_seg: Dict[int, int] = {}
    for seg_idx, seg_chunks in enumerate(segments):
        for cidx in seg_chunks:
            chunk_to_seg[cidx] = seg_idx
    
    # Map model rows to segments — a model row can span multiple segments
    seg_model_rows: Dict[int, List[int]] = {i: [] for i in range(len(segments))}
    
    for mrow in model_rows:
        chunks_hit = model_to_chunks.get(mrow.idx, [])
        if chunks_hit:
            # Find ALL segments this model row touches
            hit_segs = set()
            for cidx in chunks_hit:
                if cidx in chunk_to_seg:
                    hit_segs.add(chunk_to_seg[cidx])
            if hit_segs:
                for seg_idx in sorted(hit_segs):
                    seg_model_rows[seg_idx].append(mrow.idx)
            else:
                seg_model_rows[chunk_to_seg.get(chunks_hit[0], 0)].append(mrow.idx)
        else:
            # Fallback: assign proportionally
            ratio = mrow.idx / max(len(model_rows) - 1, 1)
            seg_idx = min(int(ratio * len(segments)), len(segments) - 1)
            seg_model_rows[seg_idx].append(mrow.idx)
    
    # Build result per segment (deduplicate model rows)
    model_row_map = {mr.idx: mr for mr in model_rows}
    result = []
    for seg_idx in range(len(segments)):
        mrow_indices = sorted(set(seg_model_rows[seg_idx]))
        merged_asr = " ".join(model_row_map[i].asr_text for i in mrow_indices if i in model_row_map)
        merged_trans = "".join(model_row_map[i].translation for i in mrow_indices if i in model_row_map)
        result.append((mrow_indices, merged_asr, merged_trans))
    
    return result


# ── 4. 主流程 ──

def align_all(
    origin_path: str,
    model_paths: Dict[str, str],  # {model_name: excel_path}
    min_segment_chars: int = 20,
    max_segment_chars: int = 500,
) -> List[AlignedSegment]:
    """
    完整对齐流程：
    1. 读原文 → 切分语义段落
    2. 读各模型 → 映射到段落
    3. 输出对齐结果
    """
    # 读取原文
    source_chunks = read_origin_txt(origin_path)
    print(f"[INFO] Source: {len(source_chunks)} chunks")
    
    # 切分语义段落
    segments = split_into_segments(source_chunks, min_segment_chars, max_segment_chars)
    print(f"[INFO] Split into {len(segments)} semantic segments")
    
    # 构建段落原文
    chunk_map = {c.idx: c for c in source_chunks}
    
    # 对齐各模型
    model_alignments = {}
    for model_name, excel_path in model_paths.items():
        model_rows = read_model_excel(excel_path)
        print(f"[INFO] {model_name}: {len(model_rows)} rows")
        aligned = align_model_to_segments(model_rows, source_chunks, segments)
        model_alignments[model_name] = aligned
    
    # 组装结果
    results = []
    for seg_idx, seg_chunks in enumerate(segments):
        source_text = " ".join(chunk_map[cidx].text for cidx in seg_chunks if cidx in chunk_map)
        
        seg = AlignedSegment(
            segment_id=seg_idx,
            source_text=source_text.strip(),
            source_chunks=seg_chunks,
            models={},
            models_asr={},
            models_rows={},
        )
        
        for model_name, aligned in model_alignments.items():
            if seg_idx < len(aligned):
                mrow_indices, merged_asr, merged_trans = aligned[seg_idx]
                seg.models[model_name] = merged_trans
                seg.models_asr[model_name] = merged_asr
                seg.models_rows[model_name] = mrow_indices
            else:
                seg.models[model_name] = ""
                seg.models_asr[model_name] = ""
                seg.models_rows[model_name] = []
        
        results.append(seg)
    
    return results


def export_to_excel(segments: List[AlignedSegment], output_path: str, model_names: List[str]):
    """导出对齐结果为 Excel"""
    rows = []
    for seg in segments:
        row = {
            "segment_id": seg.segment_id,
            "source_text": seg.source_text,
            "source_chunks": str(seg.source_chunks),
        }
        for mn in model_names:
            row[f"{mn}_asr"] = seg.models_asr.get(mn, "")
            row[f"{mn}_translation"] = seg.models.get(mn, "")
            row[f"{mn}_rows"] = str(seg.models_rows.get(mn, []))
        rows.append(row)
    
    df = pd.DataFrame(rows)
    df.to_excel(output_path, index=False, engine="openpyxl")
    print(f"[INFO] Exported {len(rows)} segments to {output_path}")


def export_to_jsonl(segments: List[AlignedSegment], output_path: str, model_names: List[str]):
    """导出对齐结果为 JSONL（可直接用于 Judge 评测）"""
    with open(output_path, "w", encoding="utf-8") as f:
        for seg in segments:
            record = {
                "asr_text": seg.source_text,
            }
            for mn in model_names:
                record[mn] = seg.models.get(mn, "")
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    print(f"[INFO] Exported {len(segments)} segments to {output_path}")


# ── CLI ──

def main():
    parser = argparse.ArgumentParser(description="同传翻译多模型分句对齐工具")
    parser.add_argument("--origin", required=True, help="人工复核的 ASR 原文 TXT 路径")
    parser.add_argument("--models", required=True, nargs="+", help="模型数据，格式: name=path.xlsx")
    parser.add_argument("--output", default="aligned_output.xlsx", help="输出文件路径 (.xlsx 或 .jsonl)")
    parser.add_argument("--min-chars", type=int, default=20, help="最小段落字符数")
    parser.add_argument("--max-chars", type=int, default=500, help="最大段落字符数")
    
    args = parser.parse_args()
    
    # Parse model args
    model_paths = {}
    for m in args.models:
        if "=" not in m:
            print(f"[ERROR] Model format should be name=path.xlsx, got: {m}")
            return
        name, path = m.split("=", 1)
        model_paths[name] = path
    
    model_names = list(model_paths.keys())
    
    # Run alignment
    segments = align_all(args.origin, model_paths, args.min_chars, args.max_chars)
    
    # Print summary
    print(f"\n{'='*60}")
    print(f"Alignment Summary:")
    print(f"  Segments: {len(segments)}")
    for mn in model_names:
        total_rows = sum(len(s.models_rows.get(mn, [])) for s in segments)
        empty_segs = sum(1 for s in segments if not s.models.get(mn, "").strip())
        print(f"  {mn}: {total_rows} rows mapped, {empty_segs} empty segments")
    print(f"{'='*60}\n")
    
    # Preview first 3 segments
    for seg in segments[:3]:
        print(f"--- Segment {seg.segment_id} ---")
        print(f"  Source: {seg.source_text[:100]}...")
        for mn in model_names:
            trans = seg.models.get(mn, "")[:80]
            print(f"  {mn}: {trans}...")
        print()
    
    # Export
    if args.output.endswith(".jsonl"):
        export_to_jsonl(segments, args.output, model_names)
    else:
        export_to_excel(segments, args.output, model_names)


if __name__ == "__main__":
    main()
